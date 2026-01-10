import io
import secrets
from datetime import datetime, date, time, timedelta

import pandas as pd
import streamlit as st
import qrcode

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build

from drive_store import load_csv, save_csv, upload_bytes, download_bytes, ensure_subfolder, list_files
from data_models import (
    PROJECTS_FILE, EMPLOYEES_FILE, REPORTS_FILE, WEEKLY_SIG_FILE,
    ensure_projects, ensure_employees, ensure_reports, ensure_weekly_signatures,
    now_utc_iso, week_id_from_date, token_hash,
    bcrypt_hash_pin, bcrypt_check_pin
)
from pdf_engine import generate_weekly_pdf
from gmail_service import send_text_email, send_pdf_email


# -------------------------
# UI / Config
# -------------------------
st.set_page_config(page_title="BauApp", layout="wide")
st.markdown(
    """
    <style>
    #MainMenu {display: none !important;}
    [data-testid="stToolbar"] {display: none !important;}
    [data-testid="stHeader"] {display: none !important;}
    footer {display: none !important;}
    </style>
    """,
    unsafe_allow_html=True,
)


def sget(key: str, default=""):
    try:
        return st.secrets.get(key, default)
    except Exception:
        return default


def require_secret(key: str) -> str:
    v = str(sget(key, "")).strip()
    if not v:
        st.error(f"Fehlendes Secret: {key}")
        st.stop()
    return v


# Secrets
GOOGLE_CLIENT_ID = require_secret("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = require_secret("GOOGLE_CLIENT_SECRET")
GOOGLE_REFRESH_TOKEN = require_secret("GOOGLE_REFRESH_TOKEN")

REPORTS_FOLDER_ID = require_secret("REPORTS_FOLDER_ID")
PHOTOS_FOLDER_ID = require_secret("PHOTOS_FOLDER_ID")
UPLOADS_FOLDER_ID = require_secret("UPLOADS_FOLDER_ID")

BASE_APP_URL = str(sget("BASE_APP_URL", "")).strip().rstrip("/")
GMAIL_SENDER = str(sget("GMAIL_SENDER", "me")).strip()
ENABLE_GMAIL = str(sget("ENABLE_GMAIL", "0")).strip() == "1"

ADMIN_PIN = str(sget("ADMIN_PIN", sget("ADMIN", ""))).strip()  # kompatibel zu deiner Benennung
EXPORTS_SUBFOLDER_NAME = str(sget("EXPORTS_SUBFOLDER_NAME", "Exports")).strip()


# -------------------------
# Google services
# -------------------------
def get_google_services():
    scopes = ["https://www.googleapis.com/auth/drive"]
    if ENABLE_GMAIL:
        scopes.append("https://www.googleapis.com/auth/gmail.send")

    creds = Credentials(
        token=None,
        refresh_token=GOOGLE_REFRESH_TOKEN,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        scopes=scopes,
    )
    if not creds.valid:
        creds.refresh(Request())

    drive = build("drive", "v3", credentials=creds)
    gmail = build("gmail", "v1", credentials=creds) if ENABLE_GMAIL else None
    return drive, gmail


drive, gmail = get_google_services()


# -------------------------
# Load + ensure tables
# -------------------------
def load_all_tables():
    df_p, pid = load_csv(drive, folder_id=REPORTS_FOLDER_ID, filename=PROJECTS_FILE)
    df_e, eid = load_csv(drive, folder_id=REPORTS_FOLDER_ID, filename=EMPLOYEES_FILE)
    df_r, rid = load_csv(drive, folder_id=REPORTS_FOLDER_ID, filename=REPORTS_FILE)
    df_s, sid = load_csv(drive, folder_id=REPORTS_FOLDER_ID, filename=WEEKLY_SIG_FILE)

    df_p = ensure_projects(df_p)
    df_e = ensure_employees(df_e)
    df_r = ensure_reports(df_r)
    df_s = ensure_weekly_signatures(df_s)
    return (df_p, pid), (df_e, eid), (df_r, rid), (df_s, sid)


def save_table(filename: str, df: pd.DataFrame, file_id):
    return save_csv(drive, folder_id=REPORTS_FOLDER_ID, filename=filename, df=df, file_id=file_id)


def compute_hours(start_t: time, end_t: time, pause_h: float) -> float:
    dt_start = datetime.combine(datetime.today(), start_t)
    dt_end = datetime.combine(datetime.today(), end_t)
    if dt_end < dt_start:
        dt_end += timedelta(days=1)
    return round(max(0.0, (dt_end - dt_start).total_seconds() / 3600 - float(pause_h or 0)), 2)


def project_qr_png(project_name: str) -> bytes:
    if BASE_APP_URL:
        link = f"{BASE_APP_URL}?project={project_name}"
    else:
        link = project_name

    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=8, border=3)
    qr.add_data(link)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


# -------------------------
# Download via token
# -------------------------
qp = st.query_params
if "download_token" in qp:
    tok = qp.get("download_token")
    (df_p, _), (df_e, _), (df_r, _), (df_s, sid) = load_all_tables()

    h = token_hash(tok)
    hit = df_s[(df_s["token_hash"] == h) & df_s["pdf_file_id"].notna()].tail(1)
    if hit.empty:
        st.error("Token ungültig oder abgelaufen.")
        st.stop()

    # expiry check (optional)
    exp = hit.iloc[0].get("token_expires_at")
    if exp:
        try:
            if datetime.utcnow() > datetime.fromisoformat(str(exp)):
                st.error("Token abgelaufen.")
                st.stop()
        except Exception:
            pass

    pdf_id = hit.iloc[0]["pdf_file_id"]
    pdf_bytes = download_bytes(drive, pdf_id)
    if not pdf_bytes:
        st.error("PDF nicht gefunden.")
        st.stop()

    st.title("Download Wochenrapport")
    st.download_button("⬇️ PDF herunterladen", pdf_bytes, file_name="Wochenrapport.pdf", mime="application/pdf")
    st.stop()


# -------------------------
# UI
# -------------------------
st.sidebar.title("BauApp")
mode = st.sidebar.radio("Bereich", ["👷 Mitarbeiter", "🛠️ Admin"], index=0)

target_project = qp.get("project", None)


# -------------------------
# Mitarbeiter
# -------------------------
if mode == "👷 Mitarbeiter":
    st.title("👷 Mitarbeiterbereich")

    (df_p, _), (df_e, _), (df_r, rid), (df_s, sid) = load_all_tables()
    df_p_active = df_p[df_p["status"] == "aktiv"].copy()

    if df_p_active.empty:
        st.warning("Keine aktiven Projekte.")
        st.stop()

    projects = df_p_active["projekt_name"].dropna().astype(str).tolist()
    default_index = projects.index(target_project) if (target_project in projects) else 0
    project_name = st.selectbox("Projekt", projects, index=default_index)

    proj = df_p_active[df_p_active["projekt_name"] == project_name].tail(1)
    projekt_id = str(proj.iloc[0]["projekt_id"]) if not proj.empty else project_name.replace(" ", "_")[:64]

    # QR
    with st.expander("📱 QR-Code für dieses Projekt"):
        png = project_qr_png(project_name)
        st.image(png, caption="Scan → Projekt in App öffnen", width=250)
        st.download_button("QR als PNG herunterladen", png, file_name=f"QR_{project_name}.png", mime="image/png")

    tab1, tab2, tab3 = st.tabs(["📝 Rapport", "🖊️ Wochen-Signatur", "📷 Fotos/📂 Pläne"])

    # -------------------------
    # Rapport
    # -------------------------
    with tab1:
        st.subheader("Rapport erfassen")

        c1, c2, c3 = st.columns(3)
        datum = c1.date_input("Datum", date.today())
        mitarbeiter = c1.text_input("Name")
        start = c2.time_input("Start", time(7, 0))
        ende = c2.time_input("Ende", time(16, 0))
        pause_h = c3.number_input("Pause (h)", 0.0, 5.0, 0.5, 0.25)

        arbeitsbeschrieb = st.text_area("Arbeitsbeschrieb / Notizen")
        material = st.text_input("Material")
        fugenfarbe = st.text_input("Fugenfarbe (optional)")
        fugen_code = st.text_input("Fugen-Code (optional)")
        asbest_relevant = st.checkbox("Asbest relevant?")
        asbest_probe = st.checkbox("Asbest Probe genommen?")

        hours = compute_hours(start, ende, pause_h)
        st.info(f"Stunden (berechnet): {hours}")

        if st.button("✅ Als DRAFT speichern", type="primary"):
            if not mitarbeiter.strip():
                st.error("Name fehlt.")
            else:
                row = {
                    "rapport_id": secrets.token_hex(16),
                    "version": 1,
                    "status": "DRAFT",
                    "created_at": now_utc_iso(),
                    "created_by": mitarbeiter.strip(),
                    "confirmed_at": None,
                    "confirmed_by": None,
                    "datum": str(datum),
                    "projekt_id": projekt_id,
                    "projekt_name": project_name,
                    "mitarbeiter_name": mitarbeiter.strip(),
                    "gast_info": None,
                    "start_time": str(start),
                    "end_time": str(ende),
                    "pause_h": float(pause_h),
                    "reisezeit_min": None,
                    "mittag": None,
                    "arbeitsbeschrieb": arbeitsbeschrieb,
                    "material": material,
                    "material_regie": None,
                    "fugenfarbe": fugenfarbe,
                    "fugen_code": fugen_code,
                    "asbest_relevant": bool(asbest_relevant),
                    "asbest_probe_genommen": bool(asbest_probe),
                    "hours": hours,
                    "correction_reason": None,
                }
                df_r = pd.concat([df_r, pd.DataFrame([row])], ignore_index=True)
                save_table(REPORTS_FILE, df_r, rid)
                st.success("Gespeichert.")
                st.rerun()

        st.divider()
        st.subheader("Rapporte (Projekt)")
        view = df_r[df_r["projekt_id"] == projekt_id].copy()
        if view.empty:
            st.info("Noch keine Rapporte.")
        else:
            view["datum"] = pd.to_datetime(view["datum"], errors="coerce")
            view = view.sort_values(["datum", "created_at"], ascending=False)
            st.dataframe(view.tail(200), use_container_width=True)

    # -------------------------
    # Wochen-Signatur
    # -------------------------
    with tab2:
        st.subheader("Wochen-Signatur")

        df_proj = df_r[(df_r["projekt_id"] == projekt_id) & df_r["datum"].notna()].copy()
        if df_proj.empty:
            st.info("Keine Rapporte vorhanden.")
            st.stop()

        df_proj["datum"] = pd.to_datetime(df_proj["datum"], errors="coerce").dt.date
        df_proj["week_id"] = df_proj["datum"].apply(week_id_from_date)

        week = st.selectbox("Woche", sorted(df_proj["week_id"].dropna().unique().tolist(), reverse=True))

        wk_rows = df_proj[df_proj["week_id"] == week].copy()
        confirmed = wk_rows[wk_rows["status"] == "CONFIRMED"].copy()

        if confirmed.empty:
            st.warning("Diese Woche ist noch nicht READY_TO_SIGN (keine CONFIRMED Rapporte).")
        else:
            st.success(f"READY_TO_SIGN: {len(confirmed)} bestätigte Rapporte")

        employee_key = st.text_input("Name (Signatur)", value="")
        pin = st.text_input("PIN", type="password")
        agree = st.checkbox("Ich bestätige die Richtigkeit der Angaben.")

        sig_hit = df_s[(df_s["week_id"] == week) & (df_s["employee_key"] == employee_key.strip())].tail(1)
        if not sig_hit.empty and sig_hit.iloc[0]["status"] == "SIGNED":
            st.info("Für diesen Namen ist die Woche bereits SIGNED.")

        if st.button("🖊️ Woche signieren", disabled=not (agree and employee_key.strip() and pin and not confirmed.empty)):
            emp = employee_key.strip()

            # optional PIN check, wenn employee in Employees.csv eingetragen ist
            rec = df_e[df_e["employee_name"].fillna("").astype(str) == emp].tail(1)
            if not rec.empty:
                h = rec.iloc[0].get("pin_hash_bcrypt")
                if h and not bcrypt_check_pin(pin, str(h)):
                    st.error("PIN falsch.")
                    st.stop()

            signature_text = f"Signiert von {emp} via CHECKBOX+PIN am {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            pdf_bytes = generate_weekly_pdf(
                company_name="R. BAUMGARTNER AG",
                week_id=week,
                employee_display=emp,
                rows_df=confirmed,
                signature_text=signature_text,
            )

            exports_folder = ensure_subfolder(drive, UPLOADS_FOLDER_ID, EXPORTS_SUBFOLDER_NAME)
            pdf_name = f"Wochenrapport_{week}_{emp}.pdf".replace(" ", "_")
            pdf_id = upload_bytes(drive, data=pdf_bytes, folder_id=exports_folder, filename=pdf_name, mimetype="application/pdf")

            token = secrets.token_urlsafe(32)
            htok = token_hash(token)

            sig_row = {
                "week_id": week,
                "employee_key": emp,
                "projekt_id": projekt_id,
                "signed_at": now_utc_iso(),
                "signed_by_display": emp,
                "signature_method": "CHECKBOX_PIN",
                "signature_image_file_id": None,
                "pdf_file_id": pdf_id,
                "status": "SIGNED",
                "invalidated_at": None,
                "invalidated_by": None,
                "invalidation_reason": None,
                "token_hash": htok,
                "token_expires_at": (datetime.utcnow() + timedelta(days=14)).isoformat(timespec="seconds"),
            }
            df_s = pd.concat([df_s, pd.DataFrame([sig_row])], ignore_index=True)
            save_table(WEEKLY_SIG_FILE, df_s, sid)

            # optional send
            sent = False
            to_email = None
            if not rec.empty:
                to_email = rec.iloc[0].get("contact_email")

            if ENABLE_GMAIL and gmail and to_email:
                try:
                    send_pdf_email(
                        gmail_service=gmail,
                        to=str(to_email),
                        subject=f"Wochenrapport {week} – Kopie",
                        body=f"Hallo {emp},\n\nanbei deine Kopie des unterschriebenen Wochenrapports ({week}).\n\nGruss\nBauApp",
                        pdf_bytes=pdf_bytes,
                        filename=pdf_name
                    )
                    sent = True
                except Exception as e:
                    st.warning(f"E-Mail konnte nicht gesendet werden: {e}")

            st.success("Signatur gespeichert und PDF archiviert.")
            if sent:
                st.success("E-Mail wurde versendet.")
            else:
                if BASE_APP_URL:
                    link = f"{BASE_APP_URL}?download_token={token}"
                    st.info("Kein Gmail-Versand (oder keine E-Mail hinterlegt). Token-Link:")
                    st.code(link)
                else:
                    st.warning("Kein BASE_APP_URL gesetzt – kein Link möglich.")

            st.download_button("⬇️ PDF jetzt herunterladen", pdf_bytes, file_name=pdf_name, mime="application/pdf")

    # -------------------------
    # Fotos / Pläne Upload
    # -------------------------
    with tab3:
        st.subheader("Fotos & Pläne")

        st.caption("Uploads werden in Drive gespeichert. Nichts wird gelöscht.")
        kind = st.radio("Typ", ["Fotos", "Pläne"], horizontal=True)

        folder = PHOTOS_FOLDER_ID if kind == "Fotos" else UPLOADS_FOLDER_ID
        subfolder = ensure_subfolder(drive, folder, project_name.replace("/", "_")[:100])

        up = st.file_uploader(f"{kind} hochladen", type=None, accept_multiple_files=True)
        if up and st.button("⬆️ Upload starten"):
            ok = 0
            for f in up:
                data = f.read()
                name = f.name
                mime = f.type or "application/octet-stream"
                upload_bytes(drive, data=data, folder_id=subfolder, filename=name, mimetype=mime)
                ok += 1
            st.success(f"{ok} Datei(en) hochgeladen.")

        st.divider()
        st.subheader("Letzte Uploads")
        files = list_files(drive, subfolder)[:30]
        if not files:
            st.info("Noch keine Dateien.")
        else:
            for f in files:
                st.write(f"• {f['name']}")

# -------------------------
# Admin
# -------------------------
else:
    st.title("🛠️ Admin")
    pin = st.text_input("Admin PIN", type="password")
    if not ADMIN_PIN or pin != ADMIN_PIN:
        st.stop()

    (df_p, pid), (df_e, eid), (df_r, rid), (df_s, sid) = load_all_tables()
    tab1, tab2, tab3, tab4 = st.tabs(["📌 Projekte", "👥 Mitarbeiter", "🧾 Rapporte", "📥 Import (Excel)"])

    # Projekte
    with tab1:
        st.subheader("Projekte")
        st.dataframe(df_p, use_container_width=True)

        c1, c2 = st.columns([0.7, 0.3])
        new_name = c1.text_input("Neues Projekt (Name)")
        new_id = c1.text_input("Projekt-ID (optional)", value="")
        if c2.button("➕ Anlegen"):
            if not new_name.strip():
                st.error("Name fehlt.")
            else:
                pid_val = new_id.strip() or new_name.strip().replace(" ", "_")[:64]
                row = {"projekt_id": pid_val, "projekt_name": new_name.strip(), "status": "aktiv"}
                df_p = pd.concat([df_p, pd.DataFrame([row])], ignore_index=True)
                save_table(PROJECTS_FILE, df_p, pid)
                st.success("Projekt angelegt.")
                st.rerun()

    # Mitarbeiter
    with tab2:
        st.subheader("Mitarbeiter")
        st.dataframe(df_e, use_container_width=True)

        with st.form("add_emp"):
            name = st.text_input("Name")
            email = st.text_input("E-Mail (für Versand)")
            rolle = st.selectbox("Rolle", ["STAMM", "TEMP_POOL"])
            status = st.selectbox("Status", ["aktiv", "inaktiv"])
            pin_plain = st.text_input("PIN (wird bcrypt-gehasht)", type="password")
            ok = st.form_submit_button("Speichern")

        if ok:
            if not name.strip():
                st.error("Name fehlt.")
            else:
                row = {
                    "employee_id": None,
                    "employee_name": name.strip(),
                    "rolle": rolle,
                    "status": status,
                    "contact_email": email.strip() or None,
                    "contact_phone": None,
                    "pin_hash_bcrypt": bcrypt_hash_pin(pin_plain) if pin_plain else None,
                }
                df_e = pd.concat([df_e, pd.DataFrame([row])], ignore_index=True)
                save_table(EMPLOYEES_FILE, df_e, eid)
                st.success("Mitarbeiter gespeichert.")
                st.rerun()

    # Rapporte confirm / correction
    with tab3:
        st.subheader("Rapporte")
        df_view = df_r.copy()
        df_view["datum"] = pd.to_datetime(df_view["datum"], errors="coerce")
        df_view = df_view.sort_values(["datum", "created_at"], ascending=False)
        st.dataframe(df_view.tail(300), use_container_width=True)

        st.divider()
        st.markdown("### ✅ DRAFT bestätigen → CONFIRMED")
        draft_ids = df_r[df_r["status"] == "DRAFT"]["rapport_id"].dropna().tolist()
        sel = st.selectbox("DRAFT rapport_id", draft_ids) if draft_ids else None
        if sel and st.button("CONFIRM"):
            m = df_r["rapport_id"] == sel
            df_r.loc[m, "status"] = "CONFIRMED"
            df_r.loc[m, "confirmed_at"] = now_utc_iso()
            df_r.loc[m, "confirmed_by"] = "ADMIN"
            save_table(REPORTS_FILE, df_r, rid)
            st.success("Bestätigt.")
            st.rerun()

        st.divider()
        st.markdown("### ✏️ Admin-Korrektur (neue Version, alte REPLACED)")
        conf_ids = df_r[df_r["status"] == "CONFIRMED"]["rapport_id"].dropna().tolist()
        sel2 = st.selectbox("CONFIRMED rapport_id", conf_ids) if conf_ids else None
        reason = st.text_input("Korrekturgrund (Pflicht)")
        if sel2 and st.button("Korrigieren"):
            if not reason.strip():
                st.error("Korrekturgrund fehlt.")
                st.stop()

            old = df_r[df_r["rapport_id"] == sel2].tail(1)
            if old.empty:
                st.error("Nicht gefunden.")
                st.stop()

            old_row = old.iloc[0].to_dict()

            # alte Version markieren
            df_r.loc[df_r["rapport_id"] == sel2, "status"] = "REPLACED"

            # neue Version erzeugen
            new_row = old_row.copy()
            new_row["rapport_id"] = secrets.token_hex(16)
            new_row["version"] = int(old_row.get("version", 1)) + 1
            new_row["status"] = "CONFIRMED"
            new_row["correction_reason"] = reason.strip()
            new_row["created_at"] = now_utc_iso()
            new_row["created_by"] = "ADMIN_CORRECTION"
            new_row["confirmed_at"] = now_utc_iso()
            new_row["confirmed_by"] = "ADMIN"

            df_r = pd.concat([df_r, pd.DataFrame([new_row])], ignore_index=True)

            # SIGNED invalidieren (falls existiert)
            try:
                d = pd.to_datetime(new_row.get("datum"), errors="coerce")
                if pd.notna(d):
                    wk = week_id_from_date(d.date())
                    emp = str(new_row.get("mitarbeiter_name") or "").strip()
                    if emp:
                        hits = df_s[(df_s["week_id"] == wk) & (df_s["employee_key"] == emp) & (df_s["status"] == "SIGNED")]
                        if not hits.empty:
                            idx = hits.index
                            df_s.loc[idx, "status"] = "SIGNED_INVALIDATED"
                            df_s.loc[idx, "invalidated_at"] = now_utc_iso()
                            df_s.loc[idx, "invalidated_by"] = "ADMIN"
                            df_s.loc[idx, "invalidation_reason"] = f"Korrektur nach Signatur: {reason.strip()}"
                            save_table(WEEKLY_SIG_FILE, df_s, sid)
            except Exception:
                pass

            save_table(REPORTS_FILE, df_r, rid)
            st.success("Korrigiert (neue Version erstellt, alte REPLACED).")
            st.rerun()

    # Import Excel
    with tab4:
        st.subheader("Import aus Excel → CSV Tabellen")
        st.caption("Importiert Projects (Aufträge) und Reports (Baustellenrapporte).")

        x_projects = st.file_uploader("Projektmanager Excel", type=["xlsx"])
        x_reports = st.file_uploader("Rapporte Excel", type=["xlsx"])

        if st.button("📥 Import starten"):
            import openpyxl

            # Projects Import
            if x_projects:
                wb = openpyxl.load_workbook(x_projects, data_only=True)
                if "Aufträge" in wb.sheetnames:
                    ws = wb["Aufträge"]
                    headers = [ws.cell(5, c).value for c in range(1, 23)]
                    rows = []
                    for r in range(6, ws.max_row + 1):
                        vals = [ws.cell(r, c).value for c in range(1, 23)]
                        if all(v is None or v == "" for v in vals):
                            continue
                        rows.append(vals)
                    dfA = pd.DataFrame(rows, columns=headers)

                    dfA = dfA[dfA["Auftragsnr."].notna() & dfA["Objekt / Auftrag"].notna()]
                    dfA["projekt_id"] = dfA["Auftragsnr."].apply(lambda v: str(int(v)) if str(v).replace(".0", "").isdigit() else str(v))
                    dfA["projekt_name"] = dfA["Objekt / Auftrag"].astype(str)
                    dfA["status"] = "aktiv"
                    df_imp = dfA[["projekt_id", "projekt_name", "status"]].drop_duplicates()

                    df_p = pd.concat([df_p, df_imp], ignore_index=True).drop_duplicates(subset=["projekt_id"], keep="last")
                    df_p = ensure_projects(df_p)
                    save_table(PROJECTS_FILE, df_p, pid)
                    st.success(f"Projects importiert: {len(df_imp)}")

            # Reports Import
            if x_reports:
                wb = openpyxl.load_workbook(x_reports, data_only=True)
                if "Baustellenrapporte" in wb.sheetnames:
                    ws = wb["Baustellenrapporte"]
                    headers = [ws.cell(4, c).value for c in range(1, 19)]
                    data = []
                    for r in range(5, ws.max_row + 1):
                        vals = [ws.cell(r, c).value for c in range(1, 19)]
                        if all(v is None or v == "" for v in vals):
                            continue
                        data.append(vals)
                    dfB = pd.DataFrame(data, columns=headers)

                    def _to_date(x):
                        if isinstance(x, datetime):
                            return x.date()
                        if isinstance(x, date):
                            return x
                        try:
                            return pd.to_datetime(x, errors="coerce").date()
                        except Exception:
                            return None

                    out_rows = []
                    for _, r in dfB.iterrows():
                        auftrag = r.get("Auftrag\nerfasst!!")
                        baustelle = r.get("Baustelle")
                        ma = r.get("MA")
                        dat = _to_date(r.get("Datum"))
                        von = r.get("Von")
                        bis = r.get("Bis")
                        pause = r.get("Pause") or r.get("Pause\nVorm. + Pause\nNachm.")
                        notiz = r.get("Notizen/Spezieles")

                        if pd.isna(auftrag) or not baustelle or not ma or not dat:
                            continue

                        pid_val = str(int(auftrag)) if str(auftrag).replace(".0", "").isdigit() else str(auftrag)
                        start_t = pd.to_datetime(str(von), errors="coerce").time() if von else time(7, 0)
                        end_t = pd.to_datetime(str(bis), errors="coerce").time() if bis else time(16, 0)
                        pause_h = float(pause) if pause not in (None, "", " ") else 0.0
                        hours = compute_hours(start_t, end_t, pause_h)

                        out_rows.append({
                            "rapport_id": secrets.token_hex(16),
                            "version": 1,
                            "status": "CONFIRMED",
                            "created_at": now_utc_iso(),
                            "created_by": "IMPORT",
                            "confirmed_at": now_utc_iso(),
                            "confirmed_by": "IMPORT",
                            "datum": str(dat),
                            "projekt_id": pid_val,
                            "projekt_name": str(baustelle),
                            "mitarbeiter_name": str(ma),
                            "gast_info": None,
                            "start_time": str(start_t),
                            "end_time": str(end_t),
                            "pause_h": pause_h,
                            "reisezeit_min": None,
                            "mittag": None,
                            "arbeitsbeschrieb": str(notiz) if notiz else "",
                            "material": "",
                            "material_regie": None,
                            "fugenfarbe": "",
                            "fugen_code": "",
                            "asbest_relevant": False,
                            "asbest_probe_genommen": False,
                            "hours": hours,
                            "correction_reason": None,
                        })

                    if out_rows:
                        df_r = pd.concat([df_r, pd.DataFrame(out_rows)], ignore_index=True)
                        df_r = ensure_reports(df_r)
                        save_table(REPORTS_FILE, df_r, rid)
                        st.success(f"Reports importiert: {len(out_rows)}")

            st.info("Import abgeschlossen.")
            st.rerun()
